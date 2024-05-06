from robocorp import workitems
from robocorp.tasks import task
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
import re
import time
from datetime import datetime, date, timedelta
import requests
import os
from openpyxl import load_workbook
import logging

@task
def work_items():
# ----------- Work Item Parameter ------------------- #
    item = workitems.inputs.current
    search_phrase= item.payload["search_phrase"]
    news_category = item.payload["news_category"]
    months = item.payload["months"]

    return search_phrase, news_category, months
    
def config_browser():
    # Configuring the webdriver path
    chrome_driver_path = r"C:\Users\Juan\Desktop\Projetos\RPA Challenge\ChromeDriver\chromedriver.exe"
    # Configurando o serviço do ChromeDriver
    service = Service(chrome_driver_path)
    # # Configurando o serviço do ChromeDriver
    driver = webdriver.Chrome(service=service)
    return driver

def open_site(driver,site):
    driver.get(site)
    driver.maximize_window()
    
def get_current_date_formatted():
    current_date_formatted = datetime.now().strftime("%m/%d/%Y")
    return current_date_formatted

def get_first_day_previous_months(num_months):
    if num_months == 0 or num_months == 1:
        num_months = 0
    else:
        num_months = num_months - 1
    current_date_time = datetime.now()
    previous_month = current_date_time.month - num_months
    year_diff = 0
    if previous_month <= 0:
        year_diff = 1
        previous_month += 12  # Ajusta para o último mês do ano anterior

    first_day_previous_month = current_date_time.replace(year=current_date_time.year - year_diff, month=previous_month, day=1)
    print(first_day_previous_month.strftime("%m/%d/%Y"))
    return first_day_previous_month.strftime("%m/%d/%Y")
    
def catch_informations(search_phrase,news_category, current_date_formatted, first_day_previous_months):

#click at the 'accept cookies' button
    setup_logging()
    logging.info("Clicking on accept Cookies button")
    element_accept_cookies = WebDriverWait(driver,30).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR,"[data-testid='Accept all-btn']"))
    )

    element_accept_cookies.click()


    # wait until the search button is visible to click
    element_search = WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR,"[data-testid='search-button']"))
    )
    element_search.click()

    logging.info("Clicking on Searching button")
    # Search the phrase in the  result field
    search_element = WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH,"//*[@id='search-input']/form/div/input"))
    )

    search_element.send_keys(search_phrase)
    logging.info("Searching the phrase:",{search_element})
    # click in the go button
    element_go = WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR,"[data-testid='search-submit']"))
    )
    element_go.click()

    #Sort by 


    element_sort_by = WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR,".css-v7it2b"))
    )
    select = Select(element_sort_by)
    select.select_by_value(news_category)

    #Click on Data Range Button
    data_range = WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR,".css-trpop8"))
    )
    logging.info("Clicking on Data Range button")
    data_range.click()

    #Click on Specific Dates
    specific_date = WebDriverWait(driver, 30).until(
    EC.element_to_be_clickable((By.XPATH,"//*[@id='site-content']/div/div[1]/div[2]/div/div/div[1]/div/div/div/ul/li[6]/button"))
    )
    logging.info("Clicking on Specific Dates button")
    specific_date.click()
        
        
    #Start Date
    start_date_field = WebDriverWait(driver, 30).until(
    EC.element_to_be_clickable((By.XPATH,"//*[@id='startDate']"))
    )
    logging.info("Start Date:", {start_date_field})
    start_date_field.send_keys(first_day_previous_months)
    #End Date
    end_date_field = WebDriverWait(driver, 30).until(
    EC.element_to_be_clickable((By.XPATH,"//*[@id='endDate']"))
    )
    logging.info("End Date:",{current_date_formatted})
    end_date_field.send_keys(current_date_formatted)

    #click in the enter button
    logging.info("Click [ENTER]")
    end_date_field.send_keys(Keys.ENTER)
    time.sleep(2)
    # click in show more to show all news
    
    selector_all_fields = "//*[@id='site-content']/div/div[2]/div[2]/ol/li[{}]"
                           
    selector_title = "//*[@id='site-content']/div/div[2]/div/ol/li[{}]/div/div/div/a/h4"
    selector_description = "//*[@id='site-content']/div/div[2]/div/ol/li[{}]/div/div/div/a/p[1]"
    selector_img = "//*[@id='site-content']/div/div[2]/div/ol/li[{}]/div/div/figure/div/img"
    selector_date = "//*[@id='site-content']/div/div[2]/div/ol/li[{}]/div/span"
    counter_lines = 0
    while True:
        counter_lines += 1
        
        #Catch the current line
        print('-------------------------------------------------------------------------------------------------')
        catch_line = driver.find_elements(By.XPATH,selector_all_fields.format(counter_lines))
        if catch_line:
            get_line = catch_line[0]
            get_line = get_line.text
        else:
            break

        #if finds advertisement skip the line                          
        find_advertisement = r"(SKIP ADVERTISEMENT|ADVERTISEMENT|Advertisement)"

        if get_line:
            if re.search(find_advertisement,get_line):
                 print(get_line)
                 print("__________________________________________________")
                 continue              
        
        catch_line = driver.find_elements(By.XPATH,selector_title.format(counter_lines))
        if catch_line:
            get_line = catch_line[0]
            title = get_line.text
            logging.info("Title of the news:", {title})
        else:
            print("O elemento NÃO foi encontrado.")
        
        catch_line = driver.find_elements(By.XPATH,selector_description.format(counter_lines))
        if catch_line:
            get_line = catch_line[0]
            description = get_line.text
            
            logging.info("Description of the news:", {description})
        else:
            print("O elemento NÃO foi encontrado.")
        
        catch_line = driver.find_elements(By.XPATH,selector_date.format(counter_lines))
        if catch_line:
            get_line = catch_line[0]
            date_field = get_line.text
            
            logging.info("Date of the news:", {date_field})
        else:
            print("O elemento NÃO foi encontrado.")



        catch_line = driver.find_elements(By.XPATH,selector_img.format(counter_lines))
       
        if catch_line:
                img_url = catch_line[0].get_attribute("src")
                print("Image URL:", img_url)
                response = requests.get(img_url)
                picture_filename = os.path.basename(img_url.split('?')[0])
                if response.status_code == 200:
                    caminho_completo = os.path.join(r'C:\Users\Juan\Desktop\Projetos\RPA Challenge\Downloads', picture_filename)
                    
                    logging.info("Image of the news:", {picture_filename})
            # Sabe the image
                    with open(caminho_completo, 'wb') as f:
                        f.write(response.content)
                else:
                    print("Error to download the image")
        else:
                
                logging.info("The picture didnt find")
                

        def count_occurrences(search_phrase, title, description):
        # Compila o padrão regex para a palavra desejada, ignorando maiúsculas/minúsculas
            pattern = re.compile(re.escape(search_phrase), re.IGNORECASE)

        # Encontra todas as correspondências no título e descrição usando o padrão regex
            title_matches = pattern.findall(title.lower())
            description_matches = pattern.findall(description.lower())

        # Retorna o número total de ocorrências encontradas
            total_matches = len(title_matches) + len(description_matches)
            return total_matches
        
        total_matches = count_occurrences(search_phrase, title, description)
        def contains_money_in_title_or_description(title, description):
    # Padrão regex para valores monetários (exemplo: $10,00 ou R$ 500.50)
            money_pattern = r'\$(\d){0,}(.|,)\d{0,}(.|,)\d{0,}|(\d){0,}(.|,)\d{0,}(.|,)\d{0,}\s*dollars|(\d){0,}(.|,)\d{0,}(.|,)\d{0,}\s*USD'

   
            title_has_money = bool(re.search(money_pattern, title))
            description_has_money = bool(re.search(money_pattern, description))

            return title_has_money or description_has_money
        
        has_money = contains_money_in_title_or_description(title, description)
        print(f"O título ou descrição contêm dinheiro: {has_money}")


        # Save informations in excel
        nome_arquivo_existente = r"C:\Users\Juan\Desktop\Projetos\RPA Challenge\Arquivos\Challenge.xlsx"
        wb = load_workbook(filename=nome_arquivo_existente)

        # Acessa a planilha desejada
        planilha = wb["Dados"]

        # Adiciona novos dados à planilha
        novos_dados = [
            [title,date_field, description, picture_filename,total_matches,has_money]
        ]
        proxima_linha = planilha.max_row + 1
        for linha in novos_dados:
            planilha.append(linha)
        wb.save(nome_arquivo_existente)

        elements_show_more = driver.find_elements(By.CSS_SELECTOR, "[data-testid='search-show-more-button']")
        if elements_show_more:
            element_show_more = elements_show_more[0]
            print("O elemento foi encontrado.")
            element_show_more.click()
        else:
            print("O elemento NÃO foi encontrado.")

        print('-------------------------------------------------------------------------------------------------------------')


        
    
    driver.quit()

try:
    # ---------- Configuration ---------- #
    site = "https://www.nytimes.com/"     
    def setup_logging(log_file="C:\\Users\\Juan\\Desktop\\Projetos\\RPA Challenge\\Logs\\automation.log", log_level=logging.INFO):
    # Configuração básica do logger
        
        log_directory = os.path.dirname(log_file)
        if not os.path.exists(log_directory):
            os.makedirs(log_directory)
        logging.basicConfig(filename=log_file, level=log_level, format="%(asctime)s - %(levelname)s - %(message)s")    


    search_phrase,news_category,months = work_items()
    driver = config_browser()
    current_date_formatted= get_current_date_formatted()
    num_months = months
    first_day_previous_months = get_first_day_previous_months(num_months)
    # ---------- Start Process ---------- #
    open_site(driver,site)

    catch_informations(search_phrase,news_category, current_date_formatted, first_day_previous_months)

except Exception as e:
    print("error:",e)
finally:
    print("Finish")
